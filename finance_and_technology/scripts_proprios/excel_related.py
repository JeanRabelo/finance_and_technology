from openpyxl import Workbook
from django.http import HttpResponse

def colocar_no_excel(historico):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',)
    response['Content-Disposition'] = 'attachment; filename=simple_xlsx_example.xlsx'
    workbook = Workbook()

    for posicionamento in historico:
        worksheet = workbook.create_sheet(title=posicionamento['data'].replace('/','-'))
        row_num = 1
        for linha in posicionamento['carteira']:
            col_num = 1
            for info in linha:
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = info
                if info.replace('-','').replace(',','').replace('.','').isnumeric():
                    cell.number_format = '#,##0.00'
                col_num = col_num + 1
            row_num = row_num + 1

    workbook.save(response)

    return response
