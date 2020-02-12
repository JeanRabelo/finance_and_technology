from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from scripts_proprios import autenticacao
from base64 import b64encode
from io import BytesIO
from openpyxl import Workbook


# Create your views here.
@login_required(login_url = '/accounts/login/')
def buscaFundo_view(request):
    [img_data, cookie_val_1, cookie_val_2] = autenticacao.imagem_e_sessao()
    imagem_captcha = b64encode(img_data).decode("utf-8")

    return render(request, 'processing/buscar_fundo.html', {'imagem_captcha' : imagem_captcha, 'cookie_val_1': cookie_val_1, 'cookie_val_2': cookie_val_2})

@login_required(login_url = '/accounts/login/')
def escolherFundo_view(request):
    if request.method == 'POST':
        enviar_dados_resultado =  autenticacao.enviar_dados(request)
        lista_fundos = enviar_dados_resultado[0]
        argumentos_aspnet = enviar_dados_resultado[1]
        cookie_val_1 = request.POST['cookie_val_1']
        cookie_val_2 = request.POST['cookie_val_2']
        return render(request, 'processing/escolher_fundo.html', {'lista_fundos' : lista_fundos, 'cookie_val_1': cookie_val_1, 'cookie_val_2': cookie_val_2, 'argumentos_aspnet': argumentos_aspnet})
    else:
        return redirect('processing:buscar_fundo')

@login_required(login_url = '/accounts/login/')
def fundo_view(request):
    if request.method == 'POST':
        [response, argumentos_aspnet, request_old] = autenticacao.retornar_fundo(request)
        table_header = response[0]
        table_content = response[1:]
        return render(request, 'processing/fundo.html', {'table_header': table_header,
        'table_content': table_content,
        'argumentos_aspnet': argumentos_aspnet,
        'cookie_val_1': request_old.POST.get('cookie_val_1'),
        'cookie_val_2': request_old.POST.get('cookie_val_2')})
    else:
        return redirect('processing:buscar_fundo')

@login_required(login_url = '/accounts/login/')
def xlsx_view(request):
    """
    Downloads all movies as Excel file with a single worksheet
    """

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',)
    response['Content-Disposition'] = 'attachment; filename=simple_xlsx_example.xlsx'
    workbook = Workbook()

    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'Summary'

    expenses = (['Aluguel', 1000], ['Gasolina',   100], ['Comida',  300], ['Academia',    50])

    # Start from the first cell. Rows and columns are zero indexed.
    row_num = 1
    col_num = 1

    # Iterate over the data and write it out row by row.
    for item, cost in (expenses):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = item
        cell = worksheet.cell(row=row_num, column=col_num+1)
        cell.value = cost
        row_num += 1

    worksheet_2 = workbook.create_sheet(title='Planilha 2',index=2)
    worksheet_2.cell(row=row_num, column=col_num).value = r'Teste de preenchimento contínuo'

    # cell = worksheet.cell(row=row_num, column=col_num)
    # worksheet.write(row, 1, '=SUM(B1:B4)')

    workbook.save(response)

    return response

    # # Define the titles for columns
    # columns = [
    #     'ID',
    #     'Title',
    #     'Description',
    #     'Length',
    #     'Rating',
    #     'Price',
    # ]
    # row_num = 1
    #
    # # Assign the titles for each cell of the header
    # for col_num, column_title in enumerate(columns, 1):
    #     cell = worksheet.cell(row=row_num, column=col_num)
    #     cell.value = column_title
    #
    # # Iterate through all movies

    # output = BytesIO()
    # workbook = xlsxwriter.Workbook(output, {'in_memory': True})
    # worksheet = workbook.add_worksheet('Summary')
    # workbook.close()
    #
    # output.seek(0)
    #
    # response = HttpResponse(output.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    # response['Content-Disposition'] = "attachment; filename=teste_para_a_microsoft.xlsx"
    #
    # output.close()
    #
    # response.write(workbook)
    # return response
